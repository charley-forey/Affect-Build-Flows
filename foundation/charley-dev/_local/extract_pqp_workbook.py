"""Extract the PQP (Project Quality Plan) workbook into versioned seed CSVs.

The client's QA/QC tracker is 44 sheets, but 26 of them are trade checklists sharing one
identical 11-column schema, and three more are statutory gate registers sharing another.
Those collapse: 44 sheets become 5 seed tables.

What is a SEED (template, same on every project, lives in git):
    qc_trades.csv           26 trades  — CSI code, DFOW ref, default risk tier
    qc_checklist_items.csv  625 items  — the actual inspection questions
    qc_gate_template.csv    102 gates  — TCO + Fire Alarm + Statutory, one shape
    qc_doh_items.csv        101 items  — NYC DOHMH pool/spa + Art.81 food service
    qc_status_vocab.csv     the workbook's data-validation lists, as a status dimension

What is NOT a seed and is deliberately not extracted here: anything a Q-Team types per
project (results, dates, inspector names, sign-offs). That is manual input, and it belongs
in SharePoint, not in git.

Re-runnable and self-checking: it asserts the counts it expects, so a workbook revision that
silently drops rows fails here rather than in a report six weeks later.

    python _local/extract_pqp_workbook.py            # dry run, prints a summary
    python _local/extract_pqp_workbook.py --apply    # write the CSVs
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
import re
import sys
from pathlib import Path

import openpyxl

REPO = Path(__file__).resolve().parents[3]
DEFAULT_WORKBOOK = REPO / "resources" / "026-025 SAUNA LOUNGE QA - QC TRACKER DRAFT 2026-7-5.xlsx"
SEED_DIR = Path(__file__).resolve().parents[1] / "02-transformation" / "seed"

# The 26 trade checklist tabs, in the order the DASHBOARD rolls them up (rows 33-58).
# DFOW ref and risk tier are read from the DASHBOARD rather than the tab, because the
# DASHBOARD is what the client's own roll-up trusts.
TRADE_TABS = [
    "Excavation", "Concrete Formwork", "Conc Reinforcement", "CIP Concrete",
    "Precast Concrete", "Unit Masonry", "Metal Deck", "Slab on Grade", "Slab on Deck",
    "Waterproofing", "Drywall Framing", "Drywall Board", "Firestopping", "Metal Frames",
    "Metal Doors", "Doors & Hardware", "Tile & Stone", "Resilient Flooring", "ACT Ceilings",
    "Millwork & Casework", "Painting", "Electrical", "Plumbing", "HVAC & Ductwork",
    "Fire Sprinkler", "Fire Alarm",
]

# Counted out of the workbook, not read off its own summary text. The COVER sheet claims
# "45 sequenced statutory gates" on the Path to TCO tab; there are 46. Trust the rows.
EXPECTED = {
    "qc_trades": 26,
    "qc_checklist_items": 625,
    "qc_gate_template": 93,      # 46 TCO + 23 Fire Alarm + 24 Statutory
    "qc_doh_items": 101,
}

# Excel silently coerces CSI section codes like "07 21 00" into dates. The workbook has
# already lost some this way. We detect rather than guess a repair — a wrong CSI code that
# looks plausible is exactly the class of defect this platform exists to catch.
CSI_RE = re.compile(r"^\d{2}[ -]?\d{2}[ -]?\d{2}$")


def slug(text: str) -> str:
    """Stable uppercase key from a sheet name: 'HVAC & Ductwork' -> 'HVAC_DUCTWORK'."""
    return re.sub(r"_+", "_", re.sub(r"[^A-Za-z0-9]+", "_", text)).strip("_").upper()


def cell(ws, row: int, col: int):
    v = ws.cell(row, col).value
    if isinstance(v, str):
        v = v.strip()
        return v or None
    return v


def clean_csi(value, where: str, defects: list[str]) -> str:
    """Return the CSI code as text, recording Excel's date coercion as a defect."""
    if value is None:
        return ""
    if isinstance(value, (dt.datetime, dt.date)):
        defects.append(f"{where}: CSI code destroyed by Excel date coercion -> {value!r}")
        return ""
    text = str(value).strip()
    if not CSI_RE.match(text.replace("  ", " ")) and text:
        # a bare number like '7' is the same coercion, one step further along
        if text.isdigit() and len(text) <= 2:
            defects.append(f"{where}: CSI code looks truncated by Excel -> {text!r}")
            return ""
    return text


def extract_trades(wb, dashboard_meta: dict, defects: list[str]) -> list[dict]:
    rows = []
    for tab in TRADE_TABS:
        ws = wb[tab]
        meta = dashboard_meta.get(tab, {})
        rows.append({
            "TradeKey": slug(tab),
            "TradeName": tab,
            "SheetName": tab,
            "CsiCode": clean_csi(cell(ws, 3, 7), f"{tab}!G3", defects),
            "DfowRef": meta.get("dfow") or cell(ws, 3, 9) or "",
            "RiskTier": meta.get("tier", ""),
            "SortOrder": len(rows) + 1,
        })
    return rows


def extract_dashboard_meta(wb, defects: list[str]) -> dict:
    """DASHBOARD rows 33-58 carry the authoritative DFOW ref and risk tier per trade."""
    ws = wb["DASHBOARD"]
    out = {}
    for r in range(33, 59):
        name = cell(ws, r, 1)
        if not name:
            continue
        tier = cell(ws, r, 3)
        out[name] = {"dfow": cell(ws, r, 2) or "", "tier": str(tier) if tier is not None else ""}
    return out


def extract_checklist_items(wb, defects: list[str]) -> list[dict]:
    """Every trade tab: items start at row 8, column B, and run to the sign-off block."""
    rows = []
    for tab in TRADE_TABS:
        ws = wb[tab]
        key = slug(tab)
        n = 0
        for r in range(8, ws.max_row + 1):
            text = cell(ws, r, 2)
            if not text:
                continue
            num = cell(ws, r, 1)
            # the Q-Team sign-off block at the foot has no item number
            if num is None:
                continue
            n += 1
            rows.append({
                "TradeKey": key,
                "ItemNumber": int(num) if str(num).isdigit() else n,
                "ItemText": text,
                "ItemKey": f"{key}-{int(num) if str(num).isdigit() else n:03d}",
            })
        if n == 0:
            defects.append(f"{tab}: no checklist items found")
    return rows


def extract_gate_template(wb, defects: list[str]) -> list[dict]:
    """Path to TCO + Path to Fire Alarm + Statutory Inspections -> one shape.

    All three are 'a gate with an authority, a predecessor, a target and a status'. The
    workbook keeps them on three tabs because Excel has no other way to group them; a
    relational model does, so they become one table with a GateType discriminator.
    """
    rows: list[dict] = []

    # --- Path to TCO: 45 gates in 7 lettered sections -------------------------------
    ws = wb["Path to TCO"]
    section = ""
    for r in range(9, 62):
        gid, label = cell(ws, r, 1), cell(ws, r, 2)
        if not label:
            continue
        if gid is None:                       # a section banner, e.g. 'A.  PERMITS...'
            section = re.sub(r"^[A-Z]\.\s*", "", label).strip()
            continue
        rows.append({
            "GateType": "TCO",
            "GateKey": f"TCO-{gid}",
            "Step": str(gid),
            "Section": section,
            "Gate": label,
            "Authority": cell(ws, r, 3) or "",
            "Agency": cell(ws, r, 4) or "",
            "Prerequisite": cell(ws, r, 5) or "",
            "Responsible": cell(ws, r, 6) or "",
            "EvidenceRequired": cell(ws, r, 7) or "",
            "SortOrder": len(rows) + 1,
        })

    # --- Path to Fire Alarm: 23 FDNY Letter-of-Approval steps ------------------------
    ws = wb["Path to Fire Alarm"]
    for r in range(8, 31):
        gid, label = cell(ws, r, 1), cell(ws, r, 2)
        if not gid or not label:
            continue
        rows.append({
            "GateType": "FIRE_ALARM",
            "GateKey": str(gid),
            "Step": str(gid),
            "Section": "FDNY Letter of Approval",
            "Gate": label,
            "Authority": "FDNY",
            "Agency": "",
            "Prerequisite": cell(ws, r, 4) or "",
            "Responsible": cell(ws, r, 3) or "",
            "EvidenceRequired": cell(ws, r, 9) or "",
            "SortOrder": len(rows) + 1,
        })

    # --- Statutory Inspections: DOB / FDNY / DOH / DEP / utility log ------------------
    ws = wb["Statutory Inspections"]
    for r in range(7, 41):
        gid, label = cell(ws, r, 1), cell(ws, r, 2)
        if not gid or not label:
            continue
        rows.append({
            "GateType": "STATUTORY",
            "GateKey": str(gid),
            "Step": str(gid),
            "Section": cell(ws, r, 4) or "",
            "Gate": label,
            "Authority": cell(ws, r, 3) or "",
            "Agency": cell(ws, r, 3) or "",
            "Prerequisite": "",
            "Responsible": "",
            "EvidenceRequired": cell(ws, r, 11) or "",
            "LinkedTcoGate": cell(ws, r, 14) or "",
            "SortOrder": len(rows) + 1,
        })
    return rows


def extract_doh_items(wb, defects: list[str]) -> list[dict]:
    """DOH Checklist: ids are 'H-nn'; lettered rows are section banners."""
    ws = wb["DOH Checklist"]
    rows, section = [], ""
    for r in range(7, ws.max_row + 1):
        gid, label = cell(ws, r, 1), cell(ws, r, 2)
        if gid and not label:
            section = re.sub(r"^[A-Z]\.\s*", "", str(gid)).strip()
            continue
        if not gid or not str(gid).startswith("H-"):
            continue
        rows.append({
            "ItemKey": str(gid),
            "Section": section,
            "Requirement": label or "",
            "Responsibility": cell(ws, r, 3) or "",
            "AffectInterface": cell(ws, r, 4) or "",
            "EvidenceRequired": cell(ws, r, 5) or "",
            "Reference": cell(ws, r, 9) or "",
            "SortOrder": len(rows) + 1,
        })
    return rows


def extract_status_vocab(wb) -> list[dict]:
    """The workbook's data-validation lists become a status dimension.

    Typing these twice — once in SharePoint choice columns, once in DAX — is how the two
    drift. One source, generated into both.
    """
    seen: dict[str, list[str]] = {}
    for ws in wb.worksheets:
        for dv in ws.data_validations.dataValidation:
            if dv.type != "list" or not dv.formula1:
                continue
            raw = dv.formula1.strip().strip('"')
            if not raw or "!" in raw:          # a range reference, not a literal list
                continue
            seen.setdefault(raw, []).append(f"{ws.title}!{dv.sqref}")

    rows = []
    for raw, locs in sorted(seen.items()):
        values = [v.strip() for v in raw.split(",") if v.strip()]
        domain = slug(re.sub(r"[^A-Za-z0-9]", "", locs[0].split("!")[0]))[:24] or "GENERAL"
        # name the domain after the first sheet that uses it, plus its shape
        domain = f"{domain}_{len(values)}"
        for i, v in enumerate(values, 1):
            rows.append({
                "Domain": domain,
                "Code": slug(v)[:48],
                "Label": v,
                "SortOrder": i,
                "IsTerminal": str(v.lower() in {
                    "complete", "closed", "accepted", "verified", "pass", "n/a", "approved",
                }).upper(),
                "UsedBy": locs[0],
            })
    return rows


def write_csv(path: Path, rows: list[dict], apply: bool) -> None:
    if not rows:
        raise SystemExit(f"refusing to write an empty seed: {path.name}")
    fields = list({k: None for row in rows for k in row})   # union, order preserved
    if not apply:
        return
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=fields, extrasaction="ignore")
        w.writeheader()
        for row in rows:
            w.writerow({f: row.get(f, "") for f in fields})


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    ap.add_argument("--out", type=Path, default=SEED_DIR)
    ap.add_argument("--apply", action="store_true", help="write the CSVs (default: dry run)")
    args = ap.parse_args()

    if not args.workbook.exists():
        print(f"workbook not found: {args.workbook}", file=sys.stderr)
        return 2

    wb = openpyxl.load_workbook(args.workbook, data_only=False)
    defects: list[str] = []

    meta = extract_dashboard_meta(wb, defects)
    tables = {
        "qc_trades": extract_trades(wb, meta, defects),
        "qc_checklist_items": extract_checklist_items(wb, defects),
        "qc_gate_template": extract_gate_template(wb, defects),
        "qc_doh_items": extract_doh_items(wb, defects),
        "qc_status_vocab": extract_status_vocab(wb),
    }

    print(f"source: {args.workbook.name}")
    print(f"target: {args.out}")
    print()
    problems = []
    for name, rows in tables.items():
        want = EXPECTED.get(name)
        got = len(rows)
        flag = ""
        if want is not None and got != want:
            flag = f"  <-- EXPECTED {want}"
            problems.append(f"{name}: expected {want} rows, got {got}")
        print(f"  {name:22s} {got:5d} rows{flag}")
        write_csv(args.out / f"{name}.csv", rows, args.apply)

    if defects:
        print("\nworkbook defects detected (report these to the client):")
        for d in defects:
            print(f"  - {d}")

    if problems:
        print("\nFAILED:")
        for p in problems:
            print(f"  - {p}")
        return 1

    print("\n" + ("written" if args.apply else "dry run — pass --apply to write"))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
